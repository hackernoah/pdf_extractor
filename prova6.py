import requests
import sys
import json
import openpyxl
import os

filename = 'test.pdf'
output = 'test.xlsx'

headers = {
    'apikey' : '6e1199bc6688957'
}

payload = {
    'language' : 'ita',
    'filetype' : 'PDF',
    'isTable' : 'true'
}
content = None
with open('test.pdf','rb') as f:
    response = requests.post('https://api.ocr.space/parse/image', headers = headers, data=payload, files={'files':f})
    if response.status_code == 200:
        content = response.json()
    else:
        print(f"ERROR: http request returned with status code {response.status_code}")
        sys.exit()

wb = openpyxl.load_workbook(output) if os.path.isfile(output) else openpyxl.Workbook()
wb.create_sheet('table')
sheet = wb['table']
row = sheet.max_row
column = 1

for result in content['ParsedResults']:
    for line in result['TextOverlay']['Lines']:
        for word in line['Words']:
            sheet.cell(row=row, column = column).value = word['WordText']
            column = column + 1
    row = row + 1
    column = 1

wb.save(output)

