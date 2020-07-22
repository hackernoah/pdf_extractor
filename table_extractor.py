import openpyxl
from openpyxl.worksheet.merge import MergedCell
import os

path = "excel1.xlsx"
wb = openpyxl.load_workbook(path) if os.path.isfile(path) else openpyxl.Workbook()
sheet = wb["Table 1"]
content = []


vertical = sheet.max_row
horizontal = sheet.max_column

# for row in range(1,vertical):
# 	data = []
# 	merged = True if type(sheet.cell(row=row, column = 1)) == 'MergedCell' else False
# 	for col in range(1,horizontal):
# 		data.append(sheet.cell(row=row, column = col).value)
# 	if merged:
# 		content[row-1] = zip(content[row-1], data)
# 	else:
# 		data = [e for e in data if e]
# 		if len(data) > 6:
# 			content.append(data)
# for row in range(1,vertical):
#     data = []
#     merged = True if type(sheet.cell(row=row+1, column = 1)) == MergedCell else False
#     for col in range(1,horizontal):
#         if merged:
#             val1 = sheet.cell(row=row, column = col).value
#             val2 = sheet.cell(row=row+1, column = col).value
#             val = (val1,val2) if val2 and val1 else val1
#             data.append(val)
#         else:
#             val = sheet.cell(row=row, column = col).value
#             if val:
#                 data.append(val)
#     data = [e for e in data if e]
#     if len(data) > 6:
#         content.append(data)

for row in range(1,vertical):
    data = []
    table = False
    merged = True if type(sheet.cell(row=row+1, column = 1)) == MergedCell and table else False
    for col in range(1,horizontal):
        if merged:
            val1 = sheet.cell(row=row, column = col).value
            val2 = sheet.cell(row=row+1, column = col).value
            val = (val1,val2) if val2 and val1 else val1
            data.append(val)
        else:
            val = sheet.cell(row=row, column = col).value
            if val:
                data.append(val)
    data = [e for e in data if e]
    if len(data) > 6:
        content.append(data)
        if not table:
            table = True 
    

print(len(content))
with open('tables.txt','w',encoding="utf-8") as f:
    for line in content:
        f.write(' ; '.join([str(el) for el in line]).replace('\n','') + '\n\n')
