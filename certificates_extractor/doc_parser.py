import openpyxl
from openpyxl.worksheet.merge import MergedCell
import os
import json
import re
import collections
import string
from datetime import date
import csv

class CertificateParser:

    def __init__(self, mapping):
        self.table_content = []
        self.other_content = []
        self.tables = []
        self.max_row = 0
        self.tracciato = self.get_mapping(mapping)
        # self.tracciato = ["Issuer",	("Isin Code","ISIN Code"), "CFI Code", ("Underlying","Underlying Shares"),"Type of Certificate", ("Underlying ISIN","ISIN of Share"),	"Strike","Issue date", ("Expiry date","Exercise Settlement Date"), "Parity",	"Nominal Value",	("Quantity","No. of Securities issued"),"Exercise Type",	"Option Type","Exercise Lot",	"Marketing name",	"Price of Underlying",	("Reference Price","Issue Price per Security"),	"Underlying currency",	"Euro-Hedged",	"1st Barrier",	"Barrier Observation",	"2nd Strike",	"2nd Barrier",	"Autocallability",	"Observation Autocallability",	"Participation %",	"Fee %",	"Long/Short",	"Bonus/Strike%",	"Cap",	"Floor",	"Coupon",	"Protection",	"Specialist Code",	"Quote Type",	"RFE Activation",	"Denomination Currency",	"Trading Currency",	"Settlement Currency",	"Settlement System",	"Leverage Number",	"Restrike %",	"Final Valuation Date",	"Professional",	"KID web link",	"Distribution Type",	"Type of Underlying",	"ACEPI Type",	"Instrument Name",	"Minimum Lot",	"Start Trading Date",	"Last Trading Date",	"Size Obligation",	"Threshold Profile ID",	"First Semaphore",	"First Error Description",	"FISN",	"Multiplier"	"File PDF"]
        self.constants = {}
        for el in self.tracciato:
            if type(el) is not tuple:
                self.constants[el] = ''
            else:
                self.constants[el[1]] = ''
        # print('TRACCIATO:')
        # print(self.tracciato)
        # print('CONSTANTS:')
        # print(self.constants)

    def get_mapping(self,path):
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        tracciato = []
        for i in range(2,sheet.max_row):
            wanted = sheet.cell(row = i, column=1).value
            present = sheet.cell(row = i, column=2).value
            if present:
                tracciato.append((wanted,present))
            else:
                tracciato.append(wanted)
        return tracciato
    
    def extract_content(self,sheet):
        vertical = sheet.max_row
        horizontal = sheet.max_column
        isin = re.compile(r'\b([A-Z]{2})((?![A-Z]{10}\b)[A-Z0-9]{6,10})\b')
        for row in range(1,vertical):
            data = []
            table = False
            merged = True if type(sheet.cell(row=row + 1, column = 1)) == MergedCell else False
            for col in range(1,horizontal):
                if merged:
                    val1 = self.check_regex(sheet.cell(row=row, column = col).value,isin)
                    val2 = self.check_regex(sheet.cell(row=row + 1, column = col).value,isin)
                    val = (val1,val2) if val1 and val2 else val1
                    if val:
                        data.append(val)
                else:
                    val = sheet.cell(row=row, column = col).value
                    if val:
                        data.append(self.check_regex(val,isin))
            data = [e for e in data if e]
            contains = False
            if len(data) > 6:
                for elem in self.table_content:
                    if data[0].find(elem[0]) >= 0:
                        contains = True
                if not contains:
                    self.table_content.append(data)
            else:
                self.other_content.append(data)

    def get_tabular_data(self):
        indexes = []
        width = 0
        n_rows = 0
        for row in self.table_content:
            new_width = len(row)
            if new_width != width:
                self.max_row = n_rows if n_rows > self.max_row else self.max_row
                n_rows = 0
                indexes.append({})
                self.tables.append({})
                for i,el in enumerate(row):
                    # print(el)
                    el = self.clean(el)
                    if el not in self.tables[-1]:
                        self.tables[-1][el] = []
                    indexes[-1][i] = el 
                width = new_width
            else:
                # rows = []
                # if tuple in [type(e) for e in row]:
                #     row1 = []
                #     row2 = []
                #     for el in row:
                #         if type(el) is tuple:
                #             row1.append(el[0])
                #             row2.append(el[1])
                #         else:
                #             row1.append(el)
                #             row2.append(el)
                #     rows.append(row1)
                #     rows.append(row2)
                # else:
                #     rows.append(row)
                # for row in rows:
                for i,el in enumerate(row):
                    if i == 0 and el in self.tables[-1][indexes[-1][0]]:
                        break
                    else:
                        self.tables[-1][indexes[-1][i]].append((n_rows,el))
                n_rows = n_rows + 1
        for t in self.tables:
            for key in list(t.keys()):
                if not len(t[key]):
                    t.pop(key)

    def find_join(self):
        join_info = []
        tables = enumerate(self.tables)
        for i,table in tables:
            for j in range(i+1,len(list(tables))):
                for key in table:
                    for  k in self.tables[j]:
                        compare1 = self.create_compare(table[key])
                        compare2 = self.create_compare(self.tables[j][k])
                        # compare1 = [str(el).replace('\n','').replace(' ','') for n,el in table[key]]
                        # compare2 = [str(el).replace('\n','').replace(' ','') for n,el in self.tables[j][k]]
                        if not (set(compare1) - set(compare2)):
                            join_info.append((i,j,key,k))
        return join_info

    def create_compare(self, lst):
        compare = []
        for n,el in lst:
            if type(el) is tuple:
                compare.append(str(el[0]).replace('\n','').replace(' ',''))
                compare.append(str(el[1]).replace('\n','').replace(' ',''))
            else:
                compare.append(str(el).replace('\n','').replace(' ',''))
        return compare

    def join_tables(self, table1, table2, col1, col2):
        new_table = {**self.tables[table1]}
        for key in self.tables[table2]:
            new_table[key] = []
        for i,el in self.tables[table1][col1]:
            if type(el) is tuple:
                for j,e in self.tables[table2][col2]:
                    if e.replace('\n','').replace(' ','') == el[0].replace('\n','').replace(' ',''):
                        for k,elem in self.tables[table2][col2]:
                            if elem.replace('\n','').replace(' ','') == el[1].replace('\n','').replace(' ',''):
                                for key in self.tables[table2]:
                                    vals = [el for m,el in self.tables[table2][key] if m==j or m==k]
                                    new_table[key].append((i,tuple(vals)))
            else:
                for j,e in self.tables[table2][col2]:
                    if e.replace('\n','').replace(' ','') == el.replace('\n','').replace(' ',''):
                        for key in self.tables[table2]:
                            for k,elem in self.tables[table2][key]:
                                if k == j:
                                    new_table[key].append((i,elem))
        self.tables = new_table


                    
    # da individuare automaticamente le colonne e le tabelle da unire
    # def join_tables(self, col1 = "Underlying Shares", col2 = "Share Company"):
    #     new_table = {**self.tables[0]}
    #     for key in self.tables[1]:
    #         new_table[key] = []

    #     for i,el in self.tables[0][col1]:
    #         for j,e in self.tables[1][col2]:
    #             if e.replace('\n','').replace(' ','') == el.replace('\n','').replace(' ',''):
    #                 for key in self.tables[1]:
    #                     for k,elem in self.tables[1][key]:
    #                         if k == j:
    #                             new_table[key].append((i,elem))
    #     self.tables = new_table

    def save_to_file(self, path):
        wb = openpyxl.load_workbook(path) if os.path.isfile(path) else openpyxl.Workbook()
        sheet = wb.active
        max_row = 0
        for key in self.tables:
            max_row = max(max_row,len(self.tables[key]))
        for i,el in enumerate(self.tracciato):
            connection = True if type(el) is tuple else False
            col = el[0] if connection else el 
            key = el[1] if connection else el
            sheet.cell(row=1, column=i + 1 ).value = col 
            if key in self.tables:
                for k,el in self.tables[key]:
                    val = str(el[0]) + '\\' + str(el[1]) if type(el) is tuple else el
                    sheet.cell(row = k + 2, column = i + 1).value = val
            else:
                if self.constants[key]:
                    if max_row:
                        for k in range(max_row):
                            sheet.cell(row = k + 2,column = i + 1 ).value = self.constants[key]
                    else:
                        sheet.cell(row = 2,column = i + 1 ).value = self.constants[key]
        time = date.today().strftime("%d/%m/%Y %H:%M:%S").replace("/",".").replace(":",".")
        issuer = self.constants[self.tracciato[0][1]]
        name = f'FinalTerm_{issuer}_{time}.csv'
        with open(f'{path + name}', 'w+', newline='') as f:
            c = csv.writer(f,delimiter=';')
            for r in sheet.rows:
                c.writerow([cell.value for cell in r])


    def create_import(self,input_path, output_path):
        wb = openpyxl.load_workbook(input_path) if os.path.isfile(input_path) else openpyxl.Workbook()
        sheet = wb.active
        self.extract_content(sheet)
        self.get_tabular_data()
        join_info = self.find_join()
        if join_info:
            for info in join_info:
                self.join_tables(*info)
        self.get_constants()
        self.save_to_file(output_path)

    def get_constants(self):
        for key in self.constants:
            for row in self.other_content:
                for el in row:
                    out = str(el).translate(str.maketrans("","", string.punctuation + '0123456789')).replace(' ','')
                    compare = key.replace(' ','')
                    if str(compare).lower() == str(out).lower():
                        self.constants[key] = row[-1] 
                        break
    

    def clean(self,txt):
        elems = txt.split()
        result = []
        stopwords = ['of','at']
        for i,el in enumerate(elems):
            if len(el) < 4 and el not in stopwords and i > 0:
                result[-1] = result[-1] + el 
            else:
                result.append(el)
        return ' '.join(result)

    def check_regex(self,val, regex):
        result = val.replace(' ','') if re.match(regex, str(val).replace(' ','')) else val
        return result


    # def reset(self):
    #     self.table_content = []
    #     self.other_content = []
    #     self.tables = []
    #     self.max_row = 0
    #     #tracciato da prendere da un file esterno e inserire inizializzazione di constants in metodo
    #     self.tracciato = ["Issuer",	("Isin Code","ISIN Code"), "CFI Code", ("Underlying","Underlying Shares"),"Type of Certificate", ("Underlying ISIN","ISIN of Share"),	"Strike","Issue date", ("Expiry date","Exercise Settlement Date"), "Parity",	"Nominal Value",	("Quantity","No. of Securities issued"),"Exercise Type",	"Option Type","Exercise Lot",	"Marketing name",	"Price of Underlying",	("Reference Price","Issue Priceper Security"),	"Underlying currency",	"Euro-Hedged",	"1st Barrier",	"Barrier Observation",	"2nd Strike",	"2nd Barrier",	"Autocallability",	"Observation Autocallability",	"Participation %",	"Fee %",	"Long/Short",	"Bonus/Strike%",	"Cap",	"Floor",	"Coupon",	"Protection",	"Specialist Code",	"Quote Type",	"RFE Activation",	"Denomination Currency",	"Trading Currency",	"Settlement Currency",	"Settlement System",	"Leverage Number",	"Restrike %",	"Final Valuation Date",	"Professional",	"KID web link",	"Distribution Type",	"Type of Underlying",	"ACEPI Type",	"Instrument Name",	"Minimum Lot",	"Start Trading Date",	"Last Trading Date",	"Size Obligation",	"Threshold Profile ID",	"First Semaphore",	"First Error Description",	"FISN",	"Multiplier"	"File PDF"]
    #     self.constants = {}
    #     for el in self.tracciato:
    #         if type(el) is not tuple:
    #             self.constants[el] = ''


