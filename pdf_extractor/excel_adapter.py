
import openpyxl
from collections import OrderedDict
import os.path

def read_concepts(path = "concetti.xlsx", sheet_name= 'concepts'):
    le = openpyxl.load_workbook(path, data_only=True)
    sheet = le.get_sheet_by_name(sheet_name)
    concepts = [(sheet.cell(row = i, column = 1).value, (sheet.cell(row = i, column = 2).value,sheet.cell(row = i, column = 3).value)) for i in range(1,sheet.max_row+1) if sheet.cell(row = i, column = 1).value ]
    return OrderedDict(concepts)

def export_values(content, path = 'export.xlsx', sheet_name = 'import'):
    wb = openpyxl.load_workbook(path) if os.path.isfile(path) else openpyxl.Workbook()
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name, 0)
    sheet = wb[sheet_name]
    progress = sheet.max_row
    if progress == 1:
        for i,key in enumerate(content.keys()):
            sheet.cell(row = progress, column = i+1).value = key
    for i,key in enumerate(content.keys()):
        sheet.cell(row = progress + 1, column = i+1).value = content[key] 
    wb.save(path)
